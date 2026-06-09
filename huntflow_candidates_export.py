#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import json
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


API_BASE = "https://api.huntflow.ru/v2"
DEFAULT_OUTPUT_DIR = "outputs"
DEFAULT_TOKEN_FILE = "secrets/huntflow_token"

RAW_HEADERS = [
    "ФИО",
    "Последнее место работы",
    "Должность",
    "Зарплата",
    "Дата рождения",
    "Текущий этап подбора",
    "Название вакансии",
    "Грейд",
    "Отдел",
    "Подразделение",
    "Дата выгрузки",
]

ANALYTICS_RENAMES = {
    "Отдел": "Цех",
    "Подразделение": "Подцех",
    "Дата выгрузки": "Дата",
}

STATUS_ENDPOINTS = [
    "/accounts/{account_id}/vacancy_statuses",
    "/accounts/{account_id}/vacancies/statuses",
    "/accounts/{account_id}/vacancy/statuses",
    "/accounts/{account_id}/statuses",
]


def now_local_date():
    return dt.datetime.now().date().isoformat()


def read_token(path):
    token = os.environ.get("HUNTFLOW_ACCESS_TOKEN")
    if token:
        return token.strip()
    token_path = Path(path)
    if token_path.exists():
        return token_path.read_text(encoding="utf-8").strip()
    raise RuntimeError(
        "Set HUNTFLOW_ACCESS_TOKEN or save the token to secrets/huntflow_token."
    )


def request_json(path, token, params=None):
    url = f"{API_BASE}{path}"
    if params:
        url = f"{url}?{urllib.parse.urlencode(params, doseq=True)}"

    req = urllib.request.Request(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "User-Agent": "hh-salary-analytics-huntflow-export/0.1",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as response:
            body = response.read().decode("utf-8")
            return json.loads(body)
    except urllib.error.HTTPError as error:
        details = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Huntflow API returned {error.code}: {details}") from error
    except urllib.error.URLError as error:
        raise RuntimeError(f"Could not reach Huntflow API: {error.reason}") from error


def iter_pages(path, token, params=None, per_page=100, delay=0.2):
    page = 1
    base_params = dict(params or {})
    while True:
        request_params = dict(base_params)
        request_params.update({"page": page, "count": per_page})
        data = request_json(path, token, request_params)
        items = extract_items(data)
        for item in items:
            yield item

        total_pages = (
            data.get("total_pages")
            or data.get("pages")
            or data.get("pagination", {}).get("total_pages")
        )
        if total_pages and page >= int(total_pages):
            break
        if not items or len(items) < per_page:
            break
        page += 1
        time.sleep(delay)


def extract_items(data):
    if isinstance(data, list):
        return data
    for key in ("items", "data", "results"):
        value = data.get(key)
        if isinstance(value, list):
            return value
    return []


def choose_account(token, account_id):
    if account_id:
        return account_id
    data = request_json("/accounts", token)
    accounts = extract_items(data)
    if not accounts:
        raise RuntimeError("No Huntflow accounts are available for this token.")
    return str(accounts[0].get("id"))


def collect_statuses(token, account_id):
    statuses = {}
    try:
        data = request_json(f"/accounts/{account_id}/vacancies/status_groups", token)
        for group in extract_items(data):
            group_name = group.get("name")
            for item in group.get("statuses") or []:
                status_id = item.get("account_vacancy_status") or item.get("status")
                if status_id and group_name:
                    statuses[str(status_id)] = str(group_name)
    except RuntimeError:
        pass

    for template in STATUS_ENDPOINTS:
        path = template.format(account_id=account_id)
        try:
            data = request_json(path, token)
        except RuntimeError:
            continue
        for item in flatten_records(extract_items(data) or [data]):
            item_id = str(item.get("id") or item.get("status") or "")
            name = item.get("name") or item.get("title") or item.get("status_name")
            if item_id and name:
                statuses[item_id] = str(name)
    return statuses


def flatten_records(value):
    if isinstance(value, dict):
        yield value
        for nested in value.values():
            yield from flatten_records(nested)
    elif isinstance(value, list):
        for item in value:
            yield from flatten_records(item)


def full_name(applicant):
    parts = [
        applicant.get("last_name") or applicant.get("lastname"),
        applicant.get("first_name") or applicant.get("firstname"),
        applicant.get("middle_name") or applicant.get("middlename"),
    ]
    name = " ".join(part.strip() for part in parts if isinstance(part, str) and part.strip())
    return name or applicant.get("name") or applicant.get("fio") or ""


def salary_text(applicant):
    if applicant.get("money"):
        return str(applicant.get("money"))
    salary = applicant.get("salary")
    if not isinstance(salary, dict):
        return "" if salary is None else str(salary)
    amount = salary.get("amount") or salary.get("money") or salary.get("value")
    currency = salary.get("currency") or salary.get("currency_code")
    if isinstance(currency, dict):
        currency = currency.get("code") or currency.get("name")
    return " ".join(str(part) for part in (amount, currency) if part not in (None, ""))


def latest_position(applicant):
    if applicant.get("company") or applicant.get("position"):
        return str(applicant.get("company") or ""), str(applicant.get("position") or "")

    workplaces = (
        applicant.get("experience")
        or applicant.get("workplaces")
        or applicant.get("last_workplace")
        or applicant.get("positions")
    )
    if isinstance(workplaces, dict):
        workplace = workplaces
    elif isinstance(workplaces, list) and workplaces:
        workplace = workplaces[0]
    else:
        workplace = {}

    if isinstance(workplace, dict):
        company = (
            workplace.get("company")
            or workplace.get("company_name")
            or workplace.get("organization")
            or workplace.get("name")
            or ""
        )
        position = (
            workplace.get("position")
            or workplace.get("post")
            or workplace.get("title")
            or workplace.get("occupation")
            or ""
        )
        return str(company or ""), str(position or "")
    return str(workplace), ""


def current_status(applicant, vacancy_id, status_names):
    links = applicant.get("links") or applicant.get("vacancies") or []
    if isinstance(links, dict):
        links = [links]
    for link in links:
        link_vacancy = link.get("vacancy") or link.get("vacancy_id")
        if isinstance(link_vacancy, dict):
            link_vacancy = link_vacancy.get("id")
        if str(link_vacancy) != str(vacancy_id):
            continue
        status = link.get("status") or link.get("vacancy_status")
        if isinstance(status, dict):
            return status.get("name") or status.get("title") or str(status.get("id") or "")
        return status_names.get(str(status), str(status or ""))

    for key in ("status", "vacancy_status"):
        status = applicant.get(key)
        if isinstance(status, dict):
            return status.get("name") or status.get("title") or str(status.get("id") or "")
        if status:
            return status_names.get(str(status), str(status))
    return ""


def find_named_field(data, aliases):
    normalized_aliases = [normalize_text(alias) for alias in aliases if alias]
    for item in flatten_records(data):
        for key, value in item.items():
            if key in {"values", "items", "fields"} and isinstance(value, list):
                found = find_named_field(value, aliases)
                if found:
                    return found
            label = ""
            if key in {"name", "title", "label"}:
                label = value
                field_value = (
                    item.get("value")
                    or item.get("text")
                    or item.get("display_value")
                    or item.get("selected")
                )
                if matches_alias(label, normalized_aliases) and field_value not in (None, ""):
                    return stringify_value(field_value)
            elif matches_alias(key, normalized_aliases) and value not in (None, "", [], {}):
                return stringify_value(value)
    return ""


def normalize_text(value):
    return str(value).strip().lower().replace("ё", "е")


def matches_alias(value, normalized_aliases):
    text = normalize_text(value)
    return any(alias == text or alias in text for alias in normalized_aliases)


def stringify_value(value):
    if isinstance(value, dict):
        return str(value.get("name") or value.get("title") or value.get("value") or value.get("id") or "")
    if isinstance(value, list):
        return ", ".join(stringify_value(item) for item in value if stringify_value(item))
    return str(value)


def vacancy_department(vacancy, aliases):
    value = find_named_field(vacancy, aliases)
    if value:
        return value
    if vacancy.get("company"):
        return str(vacancy.get("company"))
    division = vacancy.get("account_division") or vacancy.get("division")
    if isinstance(division, dict):
        return str(division.get("name") or division.get("title") or division.get("id") or "")
    return str(division or "")


def build_raw_row(applicant, vacancy, vacancy_id, status_names, export_date, args):
    company, position = latest_position(applicant)
    return {
        "ФИО": full_name(applicant),
        "Последнее место работы": company,
        "Должность": position,
        "Зарплата": salary_text(applicant),
        "Дата рождения": applicant.get("birthday") or applicant.get("birth_date") or "",
        "Текущий этап подбора": current_status(applicant, vacancy_id, status_names),
        "Название вакансии": vacancy.get("position") or vacancy.get("name") or vacancy.get("title") or "",
        "Грейд": find_named_field(vacancy, args.grade_field),
        "Отдел": vacancy_department(vacancy, args.department_field),
        "Подразделение": find_named_field(vacancy, args.subdivision_field),
        "Дата выгрузки": export_date,
    }


def to_analytics_row(row):
    result = {}
    for header in RAW_HEADERS:
        if header == "ФИО":
            continue
        result[ANALYTICS_RENAMES.get(header, header)] = row.get(header, "")
    return result


def write_csv(rows, headers, path):
    with open(path, "w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows, headers, path):
    if Workbook is None:
        raise RuntimeError("Install openpyxl to create XLSX files, or run with --format csv.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(row.get(header, ""))) for row in rows])
        ws.column_dimensions[get_column_letter(column_idx)].width = min(max(max_len + 2, 12), 38)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def collect(args):
    token = read_token(args.token_file)
    account_id = choose_account(token, args.account_id)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    export_date = args.export_date or now_local_date()
    status_names = collect_statuses(token, account_id)
    raw_rows = []

    vacancy_params = {}
    if args.vacancy_state:
        vacancy_params["state"] = args.vacancy_state

    vacancies = list(iter_pages(
        f"/accounts/{account_id}/vacancies",
        token,
        vacancy_params,
        per_page=args.per_page,
        delay=args.delay,
    ))

    for vacancy in vacancies:
        vacancy_id = vacancy.get("id")
        if not vacancy_id:
            continue
        applicant_params = {"vacancy_id": vacancy_id}
        if args.applicant_state:
            applicant_params["state"] = args.applicant_state
        for applicant in iter_pages(
            f"/accounts/{account_id}/applicants",
            token,
            applicant_params,
            per_page=args.per_page,
            delay=args.delay,
        ):
            raw_rows.append(build_raw_row(
                applicant,
                vacancy,
                vacancy_id,
                status_names,
                export_date,
                args,
            ))

    analytics_headers = [ANALYTICS_RENAMES.get(header, header) for header in RAW_HEADERS if header != "ФИО"]
    analytics_rows = [to_analytics_row(row) for row in raw_rows]
    suffix = export_date.replace("-", "")

    if args.format in {"xlsx", "both"}:
        raw_path = output_dir / f"huntflow_candidates_raw_{suffix}.xlsx"
        analytics_path = output_dir / f"huntflow_candidates_analytics_{suffix}.xlsx"
        write_xlsx(raw_rows, RAW_HEADERS, raw_path)
        write_xlsx(analytics_rows, analytics_headers, analytics_path)
        print(f"raw_xlsx={raw_path}")
        print(f"analytics_xlsx={analytics_path}")

    if args.format in {"csv", "both"}:
        raw_path = output_dir / f"huntflow_candidates_raw_{suffix}.csv"
        analytics_path = output_dir / f"huntflow_candidates_analytics_{suffix}.csv"
        write_csv(raw_rows, RAW_HEADERS, raw_path)
        write_csv(analytics_rows, analytics_headers, analytics_path)
        print(f"raw_csv={raw_path}")
        print(f"analytics_csv={analytics_path}")

    print(f"vacancies={len(vacancies)}")
    print(f"candidates={len(raw_rows)}")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export Huntflow candidates by vacancies and build an analytics-template copy."
    )
    parser.add_argument("--token-file", default=DEFAULT_TOKEN_FILE)
    parser.add_argument("--account-id", help="Huntflow organization/account id. Defaults to the first available account.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--format", choices=["xlsx", "csv", "both"], default="xlsx")
    parser.add_argument("--export-date", help="Date for the export column, YYYY-MM-DD. Defaults to today.")
    parser.add_argument("--per-page", type=int, default=100)
    parser.add_argument("--delay", type=float, default=0.2)
    parser.add_argument("--vacancy-state", default="", help="Optional Huntflow vacancy state filter.")
    parser.add_argument("--applicant-state", default="", help="Optional Huntflow applicant state filter.")
    parser.add_argument("--grade-field", action="append", default=["грейд", "grade"])
    parser.add_argument("--department-field", action="append", default=["отдел", "department", "цех"])
    parser.add_argument("--subdivision-field", action="append", default=["подразделение", "subdivision", "подцех"])
    args = parser.parse_args()

    if args.per_page < 1:
        parser.error("--per-page must be at least 1")
    return args


def main():
    args = parse_args()
    try:
        collect(args)
    except RuntimeError as error:
        print(error, file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
