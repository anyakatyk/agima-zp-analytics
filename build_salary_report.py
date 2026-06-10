#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_DB = "hh_salary.sqlite"
DEFAULT_METRICS_DB = "salary_metrics.sqlite"
DEFAULT_OUTPUT_DIR = "outputs"


def percentile(values, q):
    if not values:
        return None
    ordered = sorted(values)
    pos = (len(ordered) - 1) * q
    lower = int(pos)
    upper = min(lower + 1, len(ordered) - 1)
    weight = pos - lower
    return ordered[lower] * (1 - weight) + ordered[upper] * weight


def fetch_rows(db_path):
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=30)
    conn.row_factory = sqlite3.Row
    return [dict(row) for row in conn.execute(
        """
        SELECT
            found_at,
            search_query,
            area,
            salary_amount,
            salary_currency,
            age,
            resume_updated_at,
            resume_id
        FROM salary_observations
        ORDER BY found_at DESC, salary_amount DESC
        """
    )]


def fetch_metric_rows(metrics_db, snapshot_id=None):
    ensure_metric_report_columns(metrics_db)
    conn = sqlite3.connect(f"file:{metrics_db}?mode=ro", uri=True, timeout=30)
    conn.row_factory = sqlite3.Row
    if snapshot_id is None:
        row = conn.execute("SELECT max(id) FROM metric_snapshots").fetchone()
        snapshot_id = row[0] if row else None
    if snapshot_id is None:
        return None, [], [], []

    snapshot = dict(conn.execute(
        """
        SELECT
            id,
            created_at,
            period,
            trigger_type,
            role,
            source_system,
            workshop,
            sub_workshop,
            location,
            grade,
            stack,
            source_query,
            notes
        FROM metric_snapshots
        WHERE id = ?
        """,
        (snapshot_id,),
    ).fetchone())
    observations = [dict(row) for row in conn.execute(
        """
        SELECT
            found_at,
            source_query AS search_query,
            location AS area,
            salary_amount,
            salary_currency,
            age,
            age_bucket,
            resume_updated_at,
            role,
            source_system,
            workshop,
            sub_workshop,
            grade,
            stack,
            candidate_title,
            total_experience_months,
            total_experience_years,
            employment_form,
            viewed,
            favorited,
            marked,
            llm_role,
            llm_grade,
            llm_stack,
            llm_confidence,
            llm_reason,
            llm_enriched_at,
            llm_model,
            resume_id
        FROM salary_metric_observations
        WHERE snapshot_id = ?
        ORDER BY salary_amount DESC
        """,
        (snapshot_id,),
    )]
    groups = [dict(row) for row in conn.execute(
        """
        SELECT
            group_type,
            group_value,
            observations_count,
            salary_min,
            salary_p25,
            salary_median,
            salary_avg,
            salary_p75,
            salary_max
        FROM salary_metric_groups
        WHERE snapshot_id = ?
        ORDER BY group_type, group_value
        """,
        (snapshot_id,),
    )]
    history = [dict(row) for row in conn.execute(
        """
        SELECT
            s.id AS snapshot_id,
            s.created_at,
            s.period,
            s.trigger_type,
            s.role,
            s.source_system,
            s.location,
            s.grade,
            s.stack,
            g.observations_count,
            g.salary_median,
            g.salary_avg,
            g.salary_min,
            g.salary_max
        FROM metric_snapshots s
        LEFT JOIN salary_metric_groups g
            ON g.snapshot_id = s.id
           AND g.group_type = 'overall'
           AND g.group_value = 'all'
        ORDER BY s.id DESC
        """
    )]
    return snapshot, observations, groups, history


def ensure_metric_report_columns(metrics_db):
    conn = sqlite3.connect(metrics_db, timeout=30)
    existing_columns = {
        row[1]
        for row in conn.execute("PRAGMA table_info(salary_metric_observations)")
    }
    migrations = {
        "workshop": "ALTER TABLE salary_metric_observations ADD COLUMN workshop TEXT",
        "sub_workshop": "ALTER TABLE salary_metric_observations ADD COLUMN sub_workshop TEXT",
        "source_system": "ALTER TABLE salary_metric_observations ADD COLUMN source_system TEXT NOT NULL DEFAULT 'hh_api'",
        "llm_role": "ALTER TABLE salary_metric_observations ADD COLUMN llm_role TEXT",
        "llm_grade": "ALTER TABLE salary_metric_observations ADD COLUMN llm_grade TEXT",
        "llm_stack": "ALTER TABLE salary_metric_observations ADD COLUMN llm_stack TEXT",
        "llm_confidence": "ALTER TABLE salary_metric_observations ADD COLUMN llm_confidence REAL",
        "llm_reason": "ALTER TABLE salary_metric_observations ADD COLUMN llm_reason TEXT",
        "llm_enriched_at": "ALTER TABLE salary_metric_observations ADD COLUMN llm_enriched_at TEXT",
        "llm_model": "ALTER TABLE salary_metric_observations ADD COLUMN llm_model TEXT",
    }
    for column, sql in migrations.items():
        if column not in existing_columns:
            conn.execute(sql)
    existing_snapshot_columns = {
        row[1]
        for row in conn.execute("PRAGMA table_info(metric_snapshots)")
    }
    snapshot_migrations = {
        "workshop": "ALTER TABLE metric_snapshots ADD COLUMN workshop TEXT",
        "sub_workshop": "ALTER TABLE metric_snapshots ADD COLUMN sub_workshop TEXT",
        "source_system": "ALTER TABLE metric_snapshots ADD COLUMN source_system TEXT NOT NULL DEFAULT 'hh_api'",
    }
    for column, sql in snapshot_migrations.items():
        if column not in existing_snapshot_columns:
            conn.execute(sql)
    conn.commit()
    conn.close()


def age_bucket(age):
    if age is None:
        return "Без возраста"
    if age < 25:
        return "до 25"
    if age <= 34:
        return "25-34"
    if age <= 44:
        return "35-44"
    if age <= 54:
        return "45-54"
    return "55+"


def month_bucket(value):
    if not value:
        return "Без даты"
    return value[:7]


def numeric_stats(values):
    values = [value for value in values if value is not None]
    if not values:
        return {
            "count": 0,
            "min": None,
            "p25": None,
            "median": None,
            "avg": None,
            "p75": None,
            "max": None,
        }
    return {
        "count": len(values),
        "min": min(values),
        "p25": round(percentile(values, 0.25)),
        "median": round(percentile(values, 0.5)),
        "avg": round(sum(values) / len(values)),
        "p75": round(percentile(values, 0.75)),
        "max": max(values),
    }


def grouped_stats(rows, key_fn):
    groups = {}
    for row in rows:
        groups.setdefault(key_fn(row), []).append(row["salary_amount"])
    result = []
    for key, values in sorted(groups.items(), key=lambda item: str(item[0])):
        stats = numeric_stats(values)
        result.append([key, stats["count"], stats["min"], stats["p25"], stats["median"], stats["avg"], stats["p75"], stats["max"]])
    return result


def write_csv(rows, path):
    headers = [
        "found_at",
        "search_query",
        "area",
        "salary_amount",
        "salary_currency",
        "age",
        "resume_updated_at",
        "resume_id",
    ]
    with open(path, "w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def append_table(ws, headers, rows, start_row=1):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws, start_row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook(rows, xlsx_path):
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    salaries = [row["salary_amount"] for row in rows]
    ages = [row["age"] for row in rows if row["age"] is not None]
    stats = numeric_stats(salaries)
    age_stats = numeric_stats(ages)
    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M")

    summary["A1"] = "HH Salary Analytics"
    summary["A1"].font = Font(size=18, bold=True, color="1F4E79")
    summary["A2"] = f"Generated: {generated_at}"
    summary["A4"] = "Metric"
    summary["B4"] = "Value"
    style_header(summary, 4)
    metrics = [
        ("Records", stats["count"]),
        ("Salary min", stats["min"]),
        ("Salary p25", stats["p25"]),
        ("Salary median", stats["median"]),
        ("Salary average", stats["avg"]),
        ("Salary p75", stats["p75"]),
        ("Salary max", stats["max"]),
        ("Age median", age_stats["median"]),
    ]
    for metric in metrics:
        summary.append(metric)
    set_widths(summary, [24, 18])

    age_sheet = wb.create_sheet("By Age")
    grouped_age = grouped_stats(rows, lambda row: age_bucket(row["age"]))
    append_table(age_sheet, ["Age group", "Count", "Min", "P25", "Median", "Average", "P75", "Max"], grouped_age)
    set_widths(age_sheet, [16, 10, 12, 12, 12, 12, 12, 12])

    chart = BarChart()
    chart.title = "Median salary by age group"
    chart.y_axis.title = "RUR"
    chart.x_axis.title = "Age group"
    data = Reference(age_sheet, min_col=5, min_row=1, max_row=len(grouped_age) + 1)
    cats = Reference(age_sheet, min_col=1, min_row=2, max_row=len(grouped_age) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    age_sheet.add_chart(chart, "J2")

    updated_sheet = wb.create_sheet("By Updated Month")
    grouped_updated = grouped_stats(rows, lambda row: month_bucket(row["resume_updated_at"]))
    append_table(updated_sheet, ["Updated month", "Count", "Min", "P25", "Median", "Average", "P75", "Max"], grouped_updated)
    set_widths(updated_sheet, [18, 10, 12, 12, 12, 12, 12, 12])

    data_sheet = wb.create_sheet("Data")
    headers = ["Found at", "Search query", "Area", "Salary", "Currency", "Age", "Resume updated at", "Resume id"]
    data_rows = [[
        row["found_at"],
        row["search_query"],
        row["area"],
        row["salary_amount"],
        row["salary_currency"],
        row["age"],
        row["resume_updated_at"],
        row["resume_id"],
    ] for row in rows]
    append_table(data_sheet, headers, data_rows)
    set_widths(data_sheet, [24, 24, 10, 14, 10, 8, 24, 42])

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header in {"Salary", "Min", "P25", "Median", "Average", "P75", "Max"}:
                for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                    for item in cell:
                        item.number_format = '#,##0'

    wb.save(xlsx_path)


def build_metrics_workbook(snapshot, rows, groups, history, xlsx_path):
    wb = Workbook()
    summary = wb.active
    summary.title = "Metric Summary"

    salaries = [row["salary_amount"] for row in rows]
    ages = [row["age"] for row in rows if row["age"] is not None]
    result = numeric_stats(salaries)
    age_stats = numeric_stats(ages)

    summary["A1"] = "HH Salary Metric Snapshot"
    summary["A1"].font = Font(size=18, bold=True, color="1F4E79")
    summary["A3"] = "Field"
    summary["B3"] = "Value"
    style_header(summary, 3)
    for item in [
        ("Snapshot id", snapshot["id"]),
        ("Period", snapshot["period"]),
        ("Trigger", snapshot["trigger_type"]),
        ("Role", snapshot["role"]),
        ("Source", snapshot.get("source_system")),
        ("Workshop", snapshot.get("workshop")),
        ("Sub-workshop", snapshot.get("sub_workshop")),
        ("Location", snapshot["location"]),
        ("Grade", snapshot["grade"]),
        ("Stack", snapshot["stack"]),
        ("Source query", snapshot["source_query"]),
        ("Records", result["count"]),
        ("Salary median", result["median"]),
        ("Salary average", result["avg"]),
        ("Salary p25", result["p25"]),
        ("Salary p75", result["p75"]),
        ("Salary min", result["min"]),
        ("Salary max", result["max"]),
        ("Age median", age_stats["median"]),
        ("Notes", snapshot["notes"]),
    ]:
        summary.append(item)
    set_widths(summary, [24, 28])

    groups_sheet = wb.create_sheet("Metric Groups")
    group_rows = [[
        row["group_type"],
        row["group_value"],
        row["observations_count"],
        row["salary_min"],
        row["salary_p25"],
        row["salary_median"],
        row["salary_avg"],
        row["salary_p75"],
        row["salary_max"],
    ] for row in groups]
    append_table(
        groups_sheet,
        ["Group type", "Group value", "Count", "Min", "P25", "Median", "Average", "P75", "Max"],
        group_rows,
    )
    set_widths(groups_sheet, [22, 24, 10, 12, 12, 12, 12, 12, 12])

    history_sheet = wb.create_sheet("Snapshot History")
    history_rows = [[
        row["snapshot_id"],
        row["created_at"],
        row["period"],
        row["trigger_type"],
        row["role"],
        row.get("source_system"),
        row["location"],
        row["grade"],
        row["stack"],
        row["observations_count"],
        row["salary_median"],
        row["salary_avg"],
        row["salary_min"],
        row["salary_max"],
    ] for row in history]
    append_table(
        history_sheet,
        ["Snapshot id", "Created at", "Period", "Trigger", "Role", "Source", "Location", "Grade", "Stack", "Count", "Median", "Average", "Min", "Max"],
        history_rows,
    )
    set_widths(history_sheet, [12, 24, 12, 14, 22, 14, 18, 14, 22, 10, 12, 12, 12, 12])

    data_sheet = wb.create_sheet("Metric Data")
    data_rows = [[
        row["found_at"],
        row["role"],
        row.get("source_system"),
        row.get("workshop"),
        row.get("sub_workshop"),
        row["area"],
        row["grade"],
        row["stack"],
        row["salary_amount"],
        row["salary_currency"],
        row["age"],
        row["age_bucket"],
        row["resume_updated_at"],
        row["candidate_title"],
        row["total_experience_months"],
        row["total_experience_years"],
        row["employment_form"],
        row["viewed"],
        row["favorited"],
        row["marked"],
        row.get("llm_role"),
        row.get("llm_grade"),
        row.get("llm_stack"),
        row.get("llm_confidence"),
        row.get("llm_reason"),
        row.get("llm_enriched_at"),
        row.get("llm_model"),
        row["resume_id"],
    ] for row in rows]
    append_table(
        data_sheet,
        [
            "Found at",
            "Role",
            "Source",
            "Workshop",
            "Sub-workshop",
            "Location",
            "Grade",
            "Stack",
            "Salary",
            "Currency",
            "Age",
            "Age bucket",
            "Resume updated at",
            "Candidate title",
            "Experience months",
            "Experience years",
            "Employment form",
            "Viewed",
            "Favorited",
            "Marked",
            "LLM role",
            "LLM grade",
            "LLM stack",
            "LLM confidence",
            "LLM reason",
            "LLM enriched at",
            "LLM model",
            "Resume id",
        ],
        data_rows,
    )
    set_widths(data_sheet, [24, 22, 14, 24, 24, 18, 14, 22, 14, 10, 8, 14, 24, 32, 18, 16, 20, 10, 10, 10, 24, 14, 24, 16, 36, 24, 18, 42])

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header in {"Salary", "Min", "P25", "Median", "Average", "P75", "Max"}:
                for column_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                    for item in column_cells:
                        item.number_format = '#,##0'

    wb.save(xlsx_path)


def parse_args():
    parser = argparse.ArgumentParser(description="Build salary analytics CSV and XLSX report.")
    parser.add_argument("--db", default=DEFAULT_DB)
    parser.add_argument("--metrics-db", default=DEFAULT_METRICS_DB)
    parser.add_argument("--snapshot-id", type=int)
    parser.add_argument("--from-metrics", action="store_true", help="Build report from salary_metrics.sqlite.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main():
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if args.from_metrics:
        snapshot, rows, groups, history = fetch_metric_rows(args.metrics_db, args.snapshot_id)
        if not snapshot:
            raise SystemExit("No metric snapshots found.")
        xlsx_path = output_dir / "hh_salary_metric_report.xlsx"
        build_metrics_workbook(snapshot, rows, groups, history, xlsx_path)
        print(f"snapshot_id={snapshot['id']}")
        print(f"rows={len(rows)}")
        print(f"xlsx={xlsx_path}")
        return

    rows = fetch_rows(args.db)
    if not rows:
        raise SystemExit("No salary observations found.")

    csv_path = output_dir / "hh_salary_data.csv"
    xlsx_path = output_dir / "hh_salary_report.xlsx"
    write_csv(rows, csv_path)
    build_workbook(rows, xlsx_path)

    print(f"rows={len(rows)}")
    print(f"csv={csv_path}")
    print(f"xlsx={xlsx_path}")


if __name__ == "__main__":
    main()
